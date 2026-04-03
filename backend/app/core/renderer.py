from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from openpyxl import Workbook
from io import BytesIO
from typing import List, Dict, Any
import logging

logger = logging.getLogger(__name__)

# 注册 CJK 字体（不需要外部字体文件）
pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
CHINESE_FONT = 'STSong-Light'
logger.info("已注册 CJK 中文字体：STSong-Light")


class Renderer:
    """报表渲染器 - 生成 PDF 和 Excel"""
    
    def generate_pdf(self, data: List[Dict], columns: List[Dict], title: str = "报表") -> bytes:
        """
        生成 PDF 文件（支持中文）
        
        Args:
            data: 数据列表
            columns: 列定义 [{"field": "name", "label": "姓名"}]
            title: 报表标题
            
        Returns:
            PDF 文件字节
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # 添加标题（使用 CJK 字体支持中文）
        from reportlab.lib.styles import ParagraphStyle
        title_style = ParagraphStyle(
            name='ChineseTitle',
            parent=getSampleStyleSheet()['Heading1'],
            fontName=CHINESE_FONT,
            fontSize=18,
            leading=22,
            alignment=1  # 居中
        )
        
        elements.append(Paragraph(title, title_style))
        
        # 构建表格数据
        table_data = [
            [col.get("label", col["field"]) for col in columns],  # 表头
        ]
        
        # 添加数据行
        for row in data:
            table_data.append([str(row.get(col["field"], "")) for col in columns])
        
        # 创建表格
        table = Table(table_data, repeatRows=1)
        
        # 设置样式（使用 CJK 字体支持中文）
        style_commands = [
            # 表头样式
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), CHINESE_FONT),  # 全部使用 CJK 字体
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ]
        
        if len(data) > 1:
            style_commands.append(('BACKGROUND', (0, 2), (-1, -1), colors.lightgrey))
        
        table.setStyle(TableStyle(style_commands))
        
        elements.append(table)
        
        # 构建 PDF
        doc.build(elements)
        
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        logger.info(f"生成 PDF，{len(data)} 条记录，中文字体：{CHINESE_FONT}")
        return pdf_bytes
    
    def generate_excel(self, data: List[Dict], columns: List[Dict], title: str = "报表") -> bytes:
        """
        生成 Excel 文件（支持中文）
        
        Args:
            data: 数据列表
            columns: 列定义
            title: 报表标题
            
        Returns:
            Excel 文件字节
        """
        wb = Workbook()
        ws = wb.active
        # Excel 工作表名最长 31 字符，且不能包含特殊字符
        safe_title = title[:31].replace('/', '-').replace('\\', '-').replace('?', '-').replace('*', '-')
        ws.title = safe_title
        
        # 添加标题
        if len(columns) > 0:
            ws.merge_cells(f'A1:{self._col_letter(len(columns))}1')
        ws['A1'] = title
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
        
        # 表头
        headers = [col.get("label", col["field"]) for col in columns]
        ws.append(headers)
        
        # 数据行（确保数据类型正确）
        for row in data:
            ws.append([row.get(col["field"], "") for col in columns])
        
        # 设置表头样式
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = cell.font.copy(bold=True)
        
        # 自动调整列宽
        for col_idx, col_def in enumerate(columns, 1):
            max_length = len(str(col_def.get("label", col_def["field"])))
            for row in data[:100]:  # 只检查前 100 行
                cell_length = len(str(row.get(col_def["field"], "")))
                max_length = max(max_length, cell_length)
            ws.column_dimensions[self._col_letter(col_idx)].width = min(max_length + 2, 50)
        
        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        logger.info(f"生成 Excel，{len(data)} 条记录")
        return excel_bytes
    
    def _col_letter(self, col_index: int) -> str:
        """将列索引转换为 Excel 列字母"""
        letters = []
        while col_index > 0:
            col_index, remainder = divmod(col_index - 1, 26)
            letters.append(chr(65 + remainder))
        return ''.join(reversed(letters))
    
    def generate_excel_multi(self, all_data: list, title: str = "报表") -> bytes:
        """
        生成多组件 Excel 文件（每个组件一个工作表）
        
        Args:
            all_data: 所有组件数据 [{'title': '', 'type': '', 'columns': [], 'data': []}]
            title: 报表标题
            
        Returns:
            Excel 文件字节
        """
        wb = Workbook()
        
        # 移除默认工作表
        default_ws = wb.active
        wb.remove(default_ws)
        
        # 为每个组件创建工作表
        for i, component in enumerate(all_data):
            ws = wb.create_sheet(title=component['title'][:31])
            columns = component['columns']
            data = component['data']
            
            # 添加组件标题
            ws.merge_cells(f'A1:{self._col_letter(len(columns))}1')
            ws['A1'] = component['title']
            ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
            
            # 表头
            headers = [col.get("label", col["field"]) for col in columns]
            ws.append(headers)
            
            # 数据行
            for row in data:
                ws.append([row.get(col["field"], "") for col in columns])
            
            # 设置表头样式
            for col_idx, _ in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.font = cell.font.copy(bold=True)
            
            # 自动调整列宽
            for col_idx, col_def in enumerate(columns, 1):
                max_length = len(str(col_def.get("label", col_def["field"])))
                for row in data[:50]:
                    cell_length = len(str(row.get(col_def["field"], "")))
                    max_length = max(max_length, cell_length)
                ws.column_dimensions[self._col_letter(col_idx)].width = min(max_length + 2, 50)
        
        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        logger.info(f"生成多组件 Excel，{len(all_data)} 个组件")
        return excel_bytes
    
    def generate_pdf_multi(self, all_data: list, title: str = "报表") -> bytes:
        """
        生成多组件 PDF 文件（简化版：只生成第一个表格组件）
        
        Args:
            all_data: 所有组件数据
            title: 报表标题
            
        Returns:
            PDF 文件字节
        """
        # 找到第一个表格组件
        table_component = None
        for comp in all_data:
            if comp['type'] == 'table':
                table_component = comp
                break
        
        if not table_component:
            # 如果没有表格组件，使用第一个组件
            table_component = all_data[0]
        
        return self.generate_pdf(
            table_component['data'],
            table_component['columns'],
            f"{title} - {table_component['title']}"
        )
